#!/usr/bin/env python3
"""
generar_hoja_inbox.py — Genera una hoja nueva en el Excel de Inbox TikTok de una marca.

Flujo:
1. Lee borradores desde logs/<marca>_borradores.json (debe existir previamente,
   se genera con leer_comentarios.py + _generar_borradores_<marca>.py)
2. Encuentra el Excel en Drive sincronizado:
   <Drive>/Mi unidad/1. GESTIÓN/CUENTAS/[N. Marca]/Inbox TikTok/Inbox TikTok - [Marca].xlsx
3. Si no existe, crea el archivo + carpeta
4. Agrega una hoja nueva con timestamp YYYY-MM-DD HH:MM
5. Pobla columnas: Usuario / Tiempo / Comentario / Borrador / Acción / Categoría / Video-Link

Uso:
    python generar_hoja_inbox.py --marca manrique
    python generar_hoja_inbox.py --marca little-joe
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Falta openpyxl. Instala: pip install openpyxl")
    sys.exit(1)


SKILL_DIR = Path(__file__).resolve().parent.parent
MARCAS_FILE = SKILL_DIR / "marcas.json"
LOGS_DIR = SKILL_DIR / "logs"

COLUMNAS = ["Usuario", "Tiempo", "Comentario", "Borrador", "Acción", "Categoría", "Video / Link"]

# Colores por acción (Pattern Fill)
COLOR_RESPONDER = "C8E6C9"   # verde claro
COLOR_ESCALAR = "FFCDD2"     # rojo claro
COLOR_SKIP = "F5F5F5"        # gris claro
COLOR_HEADER = "1976D2"      # azul TikTok


def cargar_marcas():
    return json.loads(MARCAS_FILE.read_text(encoding="utf-8"))


def encontrar_borradores(marca_slug: str) -> Path:
    """Busca el JSON de borradores más reciente para la marca."""
    candidatos = [
        LOGS_DIR / f"{marca_slug}_borradores_v2.json",
        LOGS_DIR / f"{marca_slug}_borradores.json",
    ]
    for p in candidatos:
        if p.exists():
            return p
    return None


def estilo_hoja(ws, n_filas: int):
    """Aplica estilos a la hoja: header bold, colores por acción, anchos auto."""
    # Header
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, _ in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Filas de datos
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    cell_align = Alignment(vertical="top", wrap_text=True)

    for row_idx in range(2, n_filas + 2):
        accion_cell = ws.cell(row=row_idx, column=5)
        accion = (accion_cell.value or "").lower()
        if accion == "responder":
            fill = PatternFill(start_color=COLOR_RESPONDER, end_color=COLOR_RESPONDER, fill_type="solid")
        elif accion == "escalar":
            fill = PatternFill(start_color=COLOR_ESCALAR, end_color=COLOR_ESCALAR, fill_type="solid")
        else:
            fill = PatternFill(start_color=COLOR_SKIP, end_color=COLOR_SKIP, fill_type="solid")

        for col_idx in range(1, len(COLUMNAS) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill
            c.border = border
            c.alignment = cell_align

    # Ancho de columnas
    anchos = {1: 22, 2: 14, 3: 50, 4: 60, 5: 12, 6: 18, 7: 30}
    for col_idx, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    # Filas con altura adecuada para wrap
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, n_filas + 2):
        ws.row_dimensions[row_idx].height = 48

    # Freeze panes (header)
    ws.freeze_panes = "A2"


def crear_o_actualizar_excel(xlsx_path: Path, marca_label: str, borradores: list, sheet_name: str) -> dict:
    """Crea o abre el Excel y agrega una hoja con los borradores."""
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        accion = "actualizado"
        # Si la hoja default vacía sigue (primera creación), borrarla
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            try:
                del wb["Sheet"]
            except Exception:
                pass
    else:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        # Borrar la hoja default 'Sheet'
        wb.remove(wb.active)
        accion = "creado"

    # Si ya existe una hoja con el mismo nombre, agregar sufijo
    base_name = sheet_name
    suffix = 1
    while sheet_name in wb.sheetnames:
        suffix += 1
        sheet_name = f"{base_name} ({suffix})"

    ws = wb.create_sheet(title=sheet_name)

    # Header
    for col_idx, header in enumerate(COLUMNAS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Filas
    for row_idx, b in enumerate(borradores, 2):
        ws.cell(row=row_idx, column=1, value=f"@{b.get('username','')}")
        ws.cell(row=row_idx, column=2, value=b.get("tiempo", ""))
        ws.cell(row=row_idx, column=3, value=b.get("texto_original", ""))
        ws.cell(row=row_idx, column=4, value=b.get("borrador", ""))
        ws.cell(row=row_idx, column=5, value=b.get("accion", ""))
        ws.cell(row=row_idx, column=6, value=b.get("categoria", ""))
        ws.cell(row=row_idx, column=7, value=b.get("video_url", ""))

    # Estilo
    estilo_hoja(ws, len(borradores))

    # Guardar
    wb.save(xlsx_path)
    return {
        "accion": accion,
        "sheet_name": sheet_name,
        "filas_agregadas": len(borradores),
        "path": str(xlsx_path),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--marca", required=True, help="Slug de la marca (ej. manrique)")
    parser.add_argument("--nombre-hoja", help="Nombre custom de la hoja (default: timestamp)")
    args = parser.parse_args()

    data = cargar_marcas()
    if args.marca not in data["marcas"]:
        print(f"❌ Marca '{args.marca}' no existe en marcas.json")
        sys.exit(1)

    cfg = data["marcas"][args.marca]
    drive_base = data["_meta"]["drive_base_path"]
    inbox_sub = data["_meta"]["inbox_subfolder"]
    drive_folder = cfg.get("drive_folder_name")

    if not drive_folder:
        print(f"❌ Marca '{args.marca}' no tiene drive_folder_name en marcas.json")
        sys.exit(1)

    # Nombre de marca legible (sin slug-ización)
    label_map = {
        "manrique": "Manrique",
        "lozano": "Muebles Lozano",
        "distribuidora-fitness": "Distribuidora Fitness",
        "little-joe": "Little Joe",
        "mil-ideas": "Mil Ideas",
        "kintu": "Kintu",
        "novalamps": "NovaLamps",
        "la-victoria": "La Victoria",
        "oral-beauty": "Oral Beauty",
    }
    marca_label = label_map.get(args.marca, args.marca.title())

    # Ruta destino
    inbox_dir = Path(drive_base) / drive_folder / inbox_sub
    xlsx_name = f"Inbox TikTok - {marca_label}.xlsx"
    xlsx_path = inbox_dir / xlsx_name

    # Cargar borradores
    borr_path = encontrar_borradores(args.marca)
    if not borr_path:
        print(f"❌ No hay borradores en logs/ para {args.marca}")
        print(f"   Genera con: python scripts/leer_comentarios.py --marca {args.marca}")
        sys.exit(1)

    bdata = json.loads(borr_path.read_text(encoding="utf-8"))
    borradores = bdata.get("borradores", [])
    if not borradores:
        print(f"❌ logs/{borr_path.name} no tiene borradores")
        sys.exit(1)

    # Nombre de la hoja: timestamp
    sheet_name = args.nombre_hoja or datetime.now().strftime("%Y-%m-%d %H-%M")

    print(f"📋 Marca: {marca_label}")
    print(f"📂 Carpeta Drive: {inbox_dir}")
    print(f"📊 Borradores a cargar: {len(borradores)}")
    print(f"🗓️  Nombre de hoja: {sheet_name}")
    print()

    result = crear_o_actualizar_excel(xlsx_path, marca_label, borradores, sheet_name)

    print(f"✅ Excel {result['accion']}")
    print(f"   Path: {result['path']}")
    print(f"   Hoja nueva: '{result['sheet_name']}' ({result['filas_agregadas']} filas)")
    print()
    print(f"🔍 Drive se sincronizará automáticamente en unos segundos.")
    print(f"   Abre el archivo desde Drive web o desde tu Mac.")


if __name__ == "__main__":
    main()
